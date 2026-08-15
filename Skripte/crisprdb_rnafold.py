"""
CRISPRCasdb Scraper → ViennaRNA RNAfold → Excel
────────────────────────────────────────────────
Fixes v4:
  - Browser wird alle BROWSER_RESTART_EVERY Stämme neu gestartet
    (verhindert Chrome-Absturz durch RAM-Überlauf)
  - Bereits verarbeitete Stämme werden aus _backup_spacer.xlsx gelesen
    → Abgebrochener Lauf kann einfach neu gestartet werden, macht weiter
  - Backup wird nach jedem Stamm (nicht nur alle 10) gespeichert


Beim Start: CRISPR level 4 manuell im Browser setzen → Enter drücken.
Bei erneutem Start nach Absturz: Backup wird automatisch geladen.
"""

import sys, re, time, os
import RNA
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, StaleElementReferenceException,
    InvalidSessionIdException, WebDriverException
)

# ── Konfiguration ─────────────────────────────────────────────────────────────
STRAIN_URL            = "https://crisprcas.i2bc.paris-saclay.fr/MainDb/StrainList"
SCRIPT_DIR            = os.path.dirname(os.path.abspath(__file__))
OUT_FILE              = os.path.join(SCRIPT_DIR, "pseudomonas_spacer_rnafold.xlsx")
BACKUP                = os.path.join(SCRIPT_DIR, "_backup_spacer.xlsx")
PAUSE                 = 1.2
BROWSER_RESTART_EVERY = 30   # Browser alle N Stämme neu starten (bei mir nach ca. 40-60 sonst abgestürzt)
# ─────────────────────────────────────────────────────────────────────────────


def init_driver():
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1800,1000")
    opts.add_argument("--disable-gpu")
    # Speicher-Limits für Chrome setzen
    opts.add_argument("--js-flags=--max-old-space-size=512")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-infobars")
    return webdriver.Chrome(options=opts)


def js_click(driver, el):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.2)
    driver.execute_script("arguments[0].click();", el)


# ══════════════════════════════════════════════════════════════════════════════
#  Browser-Session aufsetzen (Seite laden + Filter)
# ══════════════════════════════════════════════════════════════════════════════

def setup_session(driver):
    """Lädt die Seite und wartet auf manuelle Filter-Eingabe."""
    print("  Lade StrainList …")
    driver.get(STRAIN_URL)
    time.sleep(4)

    driver.find_element(By.ID, "db-list-filter").send_keys("Pseudomonas")
    time.sleep(2)

    print("\n  → Bitte CRISPR level 4 im Browser-Fenster manuell auswählen,")
    print("    dann hier Enter drücken …")
    input()
    time.sleep(2)

    try:
        footer = driver.find_element(By.CSS_SELECTOR, "#strain-dt_info, .dataTables_info")
        print(f"  Tabelle: {footer.text}")
    except NoSuchElementException:
        pass


# ══════════════════════════════════════════════════════════════════════════════
#  Alle Stamm-Namen laden (DataTables Scroller durchscrollen)
# ══════════════════════════════════════════════════════════════════════════════

def load_all_strain_names(driver) -> list[str]:
    print("\n  Lade alle Stamm-Namen …")
    scroller = None
    for sel in [".dataTables_scrollBody", "#strain-dt_wrapper .dataTables_scrollBody"]:
        try:
            scroller = driver.find_element(By.CSS_SELECTOR, sel)
            break
        except NoSuchElementException:
            pass

    seen, names, no_new = set(), [], 0
    while no_new < 5:
        rows = driver.find_elements(By.CSS_SELECTOR, "#strain-dt tbody tr")
        new_found = False
        for row in rows:
            try:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) < 2:
                    continue
                name = cells[1].text.strip()
                if name and name not in seen:
                    seen.add(name)
                    names.append(name)
                    new_found = True
            except StaleElementReferenceException:
                continue
        no_new = 0 if new_found else no_new + 1
        if scroller:
            driver.execute_script("arguments[0].scrollTop += 2000;", scroller)
        elif rows:
            driver.execute_script("arguments[0].scrollIntoView();", rows[-1])
        time.sleep(0.7)
        print(f"    {len(names)} Namen …", end="\r")

    print(f"\n  ✓ {len(names)} Stämme gesamt.\n")
    return names


# ══════════════════════════════════════════════════════════════════════════════
#  Stamm anklicken
# ══════════════════════════════════════════════════════════════════════════════

def click_strain(driver, name: str) -> bool:
    scroller = None
    for sel in [".dataTables_scrollBody", "#strain-dt_wrapper .dataTables_scrollBody"]:
        try:
            scroller = driver.find_element(By.CSS_SELECTOR, sel)
            break
        except NoSuchElementException:
            pass

    # Im sichtbaren DOM suchen
    rows = driver.find_elements(By.CSS_SELECTOR, "#strain-dt tbody tr")
    for row in rows:
        try:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 2 and cells[1].text.strip() == name:
                js_click(driver, cells[1].find_element(By.TAG_NAME, "a"))
                time.sleep(PAUSE)
                return True
        except (StaleElementReferenceException, NoSuchElementException):
            continue

    # Scrollen und suchen
    if scroller:
        driver.execute_script("arguments[0].scrollTop = 0;", scroller)
        time.sleep(0.5)
        for _ in range(120):
            rows = driver.find_elements(By.CSS_SELECTOR, "#strain-dt tbody tr")
            for row in rows:
                try:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 2 and cells[1].text.strip() == name:
                        js_click(driver, cells[1].find_element(By.TAG_NAME, "a"))
                        time.sleep(PAUSE)
                        return True
                except (StaleElementReferenceException, NoSuchElementException):
                    continue
            driver.execute_script("arguments[0].scrollTop += 1500;", scroller)
            time.sleep(0.4)

    print(f"  [WARNUNG] Zeile nicht gefunden: {name[:50]}")
    return False


# ══════════════════════════════════════════════════════════════════════════════
#  CRISPR-Loci aus Panel lesen
# ══════════════════════════════════════════════════════════════════════════════

def get_crispr_loci(driver) -> list[str]:
    time.sleep(PAUSE)
    loci = []
    try:
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located(
                (By.XPATH, "//a[contains(@onclick, \"'CRISPR'\")]")
            )
        )
    except TimeoutException:
        return loci

    for lnk in driver.find_elements(By.XPATH, "//a[contains(@onclick, \"'CRISPR'\")]"):
        onclick = lnk.get_attribute("onclick") or ""
        m = re.search(r"elSelected\([^,]+,\s*'CRISPR',\s*'([^']+)'", onclick)
        if m:
            lid = m.group(1)
            if lid not in loci:
                loci.append(lid)
    return loci


# ══════════════════════════════════════════════════════════════════════════════
#  Spacer aus id=spacer lesen
# ══════════════════════════════════════════════════════════════════════════════

def get_spacers_for_locus(driver, locus_id: str) -> list[str]:
    try:
        locus_link = driver.find_element(
            By.XPATH, f"//a[contains(@onclick, \"'CRISPR', '{locus_id}'\")]"
        )
        js_click(driver, locus_link)
        time.sleep(PAUSE)
    except NoSuchElementException:
        print(f"    [WARNUNG] Link nicht gefunden: {locus_id}")
        return []

    # Warte bis id=spacer Inhalt hat (max 8s, dann weitermachen)
    try:
        WebDriverWait(driver, 8).until(
            lambda d: bool(d.find_element(By.ID, "spacer").get_attribute("innerHTML").strip())
        )
    except (TimeoutException, NoSuchElementException):
        pass

    spacers = []
    try:
        html = driver.find_element(By.ID, "spacer").get_attribute("innerHTML")
        parts = re.findall(r"<div>([^<>]+)</div>", html)
        for part in parts:
            part = part.strip()
            if part.startswith(">") or part.startswith("&gt;"):
                continue
            seq = re.sub(r"[^ACGTacgt]", "", part).upper()
            if len(seq) > 5:
                spacers.append(seq)
    except NoSuchElementException:
        print(f"    [WARNUNG] id=spacer nicht gefunden: {locus_id}")

    return spacers


# ══════════════════════════════════════════════════════════════════════════════
#  Backup laden (für Fortsetzung nach Absturz)
# ══════════════════════════════════════════════════════════════════════════════

def load_backup() -> tuple[list[dict], set[str]]:
    """
    Lädt vorhandenes Backup. Gibt (collected, done_names) zurück.
    done_names = Menge der bereits vollständig verarbeiteten Organismusnamen.
    """
    if not os.path.exists(BACKUP):
        return [], set()
    try:
        df = pd.read_excel(BACKUP)
        collected = [{"Organismus": r["Organismus"], "DNA": r["Sequenz"]}
                     for _, r in df.iterrows()]
        # Als "fertig" gelten alle Organismen die im Backup stehen
        done = set(df["Organismus"].unique())
        print(f"  ✓ Backup geladen: {len(collected)} Spacer, {len(done)} Stämme bereits fertig.")
        print(f"    Setze fort ab dem nächsten noch nicht verarbeiteten Stamm.\n")
        return collected, done
    except Exception as e:
        print(f"  [Backup-Fehler] {e} – starte neu.")
        return [], set()


def save_backup(collected: list[dict]):
    pd.DataFrame([
        {"Organismus": r["Organismus"], "Sequenz": r["DNA"]}
        for r in collected
    ]).to_excel(BACKUP, index=False)


# ══════════════════════════════════════════════════════════════════════════════
#  ViennaRNA
# ══════════════════════════════════════════════════════════════════════════════

def fold(rna: str) -> float | None:
    try:
        _, mfe = RNA.fold(rna)
        return round(mfe, 2)
    except Exception as e:
        print(f"      [RNAfold] {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  Excel-Formatierung
# ══════════════════════════════════════════════════════════════════════════════

def style_excel(path: str, nrows: int):
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Spacer RNAfold"
    hfill  = PatternFill("solid", start_color="2F5496", end_color="2F5496")
    afill  = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 28
    for r in range(2, nrows + 2):
        fill = afill if r % 2 == 0 else None
        for cell in ws[r]:
            cell.font   = Font(name="Arial", size=10)
            cell.border = border
            if fill:
                cell.fill = fill
            if cell.column == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif cell.column == 2:
                cell.font      = Font(name="Courier New", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.00"
        ws.row_dimensions[r].height = 15
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 24
    ws.auto_filter.ref = f"A1:C{nrows + 1}"
    ws.freeze_panes    = "A2"
    wb.save(path)
    print(f"  ✓ Gespeichert: {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  Hauptprogramm
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("═══ CRISPRCasdb → RNAfold → Excel ═══\n")

    # Backup laden (falls vorhanden → Fortsetzung nach Absturz)
    collected, done_names = load_backup()

    driver      = None
    names       = None
    first_start = True

    try:
        # ── Erste Browser-Session ─────────────────────────────────────────────
        driver = init_driver()
        setup_session(driver)
        first_start = False

        names = load_all_strain_names(driver)
        if not names:
            print("Keine Stämme gefunden!")
            return
        total = len(names)

        # Überspringe bereits fertige Stämme
        todo = [n for n in names if n not in done_names]
        skipped = total - len(todo)
        if skipped:
            print(f"  Überspringe {skipped} bereits verarbeitete Stämme.")
        print(f"  Verbleibend: {len(todo)} Stämme.\n")

        # ── Pro Stamm: Spacer sammeln ─────────────────────────────────────────
        for s_idx, name in enumerate(todo, 1):
            print(f"[{s_idx:3d}/{len(todo)}] {name}")

            # Browser alle N Stämme neu starten
            if s_idx > 1 and (s_idx - 1) % BROWSER_RESTART_EVERY == 0:
                print(f"\n  ♻ Browser-Neustart (alle {BROWSER_RESTART_EVERY} Stämme) …")
                try:
                    driver.quit()
                except Exception:
                    pass
                time.sleep(2)
                driver = init_driver()
                setup_session(driver)
                print()

            try:
                if not click_strain(driver, name):
                    continue

                loci = get_crispr_loci(driver)
                if not loci:
                    print(f"  Keine CRISPR-Loci.")
                    # Trotzdem als "fertig" markieren
                    done_names.add(name)
                    continue
                print(f"  Loci: {loci}")

                for locus_id in loci:
                    spacers = get_spacers_for_locus(driver, locus_id)
                    print(f"  {locus_id}: {len(spacers)} Spacer")
                    for dna in spacers:
                        collected.append({"Organismus": name, "DNA": dna})

                done_names.add(name)

            except (InvalidSessionIdException, WebDriverException) as e:
                print(f"\n  [Browser-Fehler] {type(e).__name__} – starte neu …")
                save_backup(collected)
                try:
                    driver.quit()
                except Exception:
                    pass
                time.sleep(3)
                driver = init_driver()
                setup_session(driver)
                # Aktuellen Stamm nochmal versuchen
                continue

            # Nach jedem Stamm Backup speichern
            save_backup(collected)

    except KeyboardInterrupt:
        print("\n[Abbruch durch Benutzer]")
    finally:
        save_backup(collected)
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    if not collected:
        print("\nKeine Spacer gesammelt.")
        return

    # ── RNAfold ───────────────────────────────────────────────────────────────
    print(f"\nFalte {len(collected)} Spacer mit ViennaRNA …")
    rows = []
    for i, entry in enumerate(collected, 1):
        rna = entry["DNA"].replace("T", "U")
        mfe = fold(rna)
        if mfe is not None:
            rows.append({
                "Organismus":               entry["Organismus"],
                "RNA-Sequenz":              rna,
                "Freie Energie (kcal/mol)": mfe,
            })
        if i % 100 == 0 or i == len(collected):
            print(f"  {i}/{len(collected)} gefaltet …")

    # ── Sortieren & Excel ─────────────────────────────────────────────────────
    df = pd.DataFrame(rows)
    df.sort_values("Freie Energie (kcal/mol)", ascending=True, inplace=True)
    df.reset_index(drop=True, inplace=True)
    df.to_excel(OUT_FILE, index=False, engine="openpyxl")
    style_excel(OUT_FILE, len(df))

    print(f"\n{'═'*50}")
    print(f"  Fertig!  {len(df)} Spacer  →  '{OUT_FILE}'")
    print(f"  Sortiert: niedrigste → höchste freie Energie")
    print(f"{'═'*50}")


if __name__ == "__main__":
    main()
