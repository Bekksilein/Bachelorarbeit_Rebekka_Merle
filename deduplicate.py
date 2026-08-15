"""
Entfernt doppelte Sequenzen aus der RNAfold-Ergebnistabelle.
Gleiche Sequenzen werden zusammengeführt, alle Organismusnamen
werden in einer Zelle kombiniert (durch " | " getrennt).

Starten:
    py deduplicate.py

Um eine andere Eingabedatei zu verwenden:
    → Zeile 17 ändern: IN_FILE = "anderer_dateiname.xlsx"
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Hier die Eingabedatei ändern ──────────────────────────────────────────────
IN_FILE  = os.path.join(SCRIPT_DIR, "spacer_mit _32bp_rnafold.xlsx")  # <── DIESE ZEILE
# ─────────────────────────────────────────────────────────────────────────────

OUT_FILE = os.path.join(SCRIPT_DIR, "dedupliziert_" + os.path.basename(IN_FILE))

# Einlesen
df = pd.read_excel(IN_FILE)
print(f"  Eingelesen: {len(df)} Zeilen")
print(f"  Spaltennamen: {list(df.columns)}")

# Spaltennamen automatisch erkennen (flexibel gegen Tippvarianten)
def find_col(df, keywords):
    for col in df.columns:
        if any(k.lower() in col.lower() for k in keywords):
            return col
    raise ValueError(f"Keine Spalte gefunden für: {keywords}\nVorhandene Spalten: {list(df.columns)}")

org_col = find_col(df, ["organismus", "organism", "strain"])
seq_col = find_col(df, ["sequenz", "sequence", "seq", "rna"])
eng_col = find_col(df, ["energie", "energy", "mfe", "kcal"])

print(f"  Organismus-Spalte : '{org_col}'")
print(f"  Sequenz-Spalte    : '{seq_col}'")
print(f"  Energie-Spalte    : '{eng_col}'")

# Gruppieren nach Sequenz: Organismen zusammenführen, Energie beibehalten
# groupby macht seq_col zum Index → deshalb als reset_index zurückholen
grouped = df.groupby(seq_col, sort=False)

rows = []
for seq, g in grouped:
    orgs = g[org_col].dropna().unique().tolist()
    rows.append({
        org_col: " | ".join(orgs),
        seq_col: seq,
        eng_col: g[eng_col].iloc[0],
    })

result = pd.DataFrame(rows, columns=[org_col, seq_col, eng_col])
result.sort_values(eng_col, ascending=True, inplace=True)
result.reset_index(drop=True, inplace=True)

print(f"\n  Vorher : {len(df)} Zeilen")
print(f"  Nachher: {len(result)} einzigartige Sequenzen")
print(f"  Entfernt: {len(df) - len(result)} Duplikate")

# Excel speichern
result.to_excel(OUT_FILE, index=False, engine="openpyxl")

# Formatieren
wb = load_workbook(OUT_FILE)
ws = wb.active
ws.title = "Dedupliziert"
hfill  = PatternFill("solid", start_color="2F5496", end_color="2F5496")
afill  = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
thin   = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for cell in ws[1]:
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill      = hfill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border
ws.row_dimensions[1].height = 28
for r in range(2, len(result) + 2):
    fill = afill if r % 2 == 0 else None
    for cell in ws[r]:
        cell.font   = Font(name="Arial", size=10)
        cell.border = border
        if fill: cell.fill = fill
        if cell.column == 1:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        elif cell.column == 2:
            cell.font      = Font(name="Courier New", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment     = Alignment(horizontal="right", vertical="center")
            cell.number_format = "0.00"
    org_text = str(result.iloc[r-2][org_col])
    n_orgs   = org_text.count(" | ") + 1
    ws.row_dimensions[r].height = max(15, n_orgs * 14)

ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 65
ws.column_dimensions["C"].width = 24
ws.auto_filter.ref = f"A1:C{len(result)+1}"
ws.freeze_panes    = "A2"
wb.save(OUT_FILE)

print(f"\n  ✓ Gespeichert: {OUT_FILE}")
