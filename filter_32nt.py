"""
Filtert alle Spacer mit genau 32 Basen aus der Ergebnis-Tabelle
und speichert sie sortiert in eine neue Excel-Datei.

Starten:
    py filter_32bp.py
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
IN_FILE    = os.path.join(SCRIPT_DIR, "pseudomonas_spacer_rnafold.xlsx")
OUT_FILE   = os.path.join(SCRIPT_DIR, "spacer_32bp_rnafold.xlsx")


def style_excel(path: str, nrows: int):
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Spacer 32bp"
    hfill  = PatternFill("solid", start_color="2F5496", end_color="2F5496")
    afill  = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 28
    for r in range(2, nrows + 2):
        fill = afill if r % 2 == 0 else None
        for cell in ws[r]:
            cell.font   = Font(name="Arial", size=10)
            cell.border = border
            if fill:
                cell.fill = fill
            if cell.column == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif cell.column == 2:
                cell.font      = Font(name="Courier New", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment     = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.00"
        ws.row_dimensions[r].height = 15
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 24
    ws.auto_filter.ref = f"A1:C{nrows + 1}"
    ws.freeze_panes    = "A2"
    wb.save(path)


def main():
    print(f"Lese: {IN_FILE}")
    df = pd.read_excel(IN_FILE)

    seq_col = "RNA-Sequenz"
    eng_col = "Freie Energie (kcal/mol)"

    # Länge berechnen und auf 32 filtern
    df["Länge"] = df[seq_col].astype(str).str.len()
    df32 = df[df["Länge"] == 32][["Organismus", seq_col, eng_col]].copy()
    df32 = df32.sort_values(eng_col, ascending=True).reset_index(drop=True)

    print(f"Gesamt:       {len(df)} Spacer")
    print(f"Mit 32 Basen: {len(df32)} Spacer")

    df32.to_excel(OUT_FILE, index=False, engine="openpyxl")
    style_excel(OUT_FILE, len(df32))
    print(f"Gespeichert:  {OUT_FILE}")


if __name__ == "__main__":
    main()
